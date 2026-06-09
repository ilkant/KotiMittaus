"""
measurement.py — mittausten tietoluokat ja niihin liittyvät apufunktiot.

Sisältää:
  • Laitekohtaiset mittausluokat (GlucoseMeasurement, BloodPressureMeasurement,
    BodyCompositionMeasurement) — kussakin vain kyseisen laitteen kentät.
  • DEVICE_TYPES — laitelista, joka sitoo laiteavaimen sen mittausluokkaan.
  • Measurement — yhdistetty mittausrivi (yksi rivi Excel-taulukossa).
  • combine_measurements_by_day — kokoaa laitteiden lukemat päivätasolle:
    yksi Measurement per kalenteripäivä, tuorein lukema per laite.
  • write_measurements_by_date — kirjoittaa mittaukset päiväpohjaiseen
    taulukkoon (yksi rivi per päivä) oikean päivän riville. Täyttää vain
    tyhjät solut: taulukossa jo oleva arvo ei koskaan ylikirjoitu.

Uusi laite lisätään määrittelemällä sille oma mittausluokka ja rekisteröimällä
se DEVICE_TYPES-listaan. Loput koodista käsittelee laitteet listan kautta.
"""

from dataclasses import dataclass, fields
from datetime import datetime, timedelta, date
from typing import Optional


# ═══════════════════════════════════════════════════════════════════════════════
# LAITEKOHTAISET MITTAUSLUOKAT — kussakin vain yhden laitteen kentät
# ═══════════════════════════════════════════════════════════════════════════════

class DeviceMeasurement:
    """Kantaluokka laitekohtaisille mittausluokille.

    Aliluokat ovat dataclasseja, joissa on vain kyseisen laitteen tuottamat
    kentät (sekä yhteinen `timestamp`). Yhteiset apumetodit hyödyntävät
    dataclass-kenttiä, joten uusi laite tarvitsee vain oman dataclassinsa.
    """

    @classmethod
    def from_dict(cls, d: dict) -> "DeviceMeasurement":
        """Rakentaa olion lukijan tuottamasta sanakirjasta poimien vain
        tämän luokan kentät (ylimääräiset avaimet kuten 'seq'/'user' ohitetaan)."""
        names = {f.name for f in fields(cls)}
        return cls(**{k: v for k, v in d.items() if k in names})

    def merge_into(self, m: "Measurement") -> None:
        """Kopioi tämän laitteen kentät (paitsi timestamp) yhdistettyyn
        Measurement-riviin. Vain ei-tyhjät arvot ylikirjoittavat."""
        for f in fields(self):
            if f.name == "timestamp":
                continue
            value = getattr(self, f.name)
            if value is not None and hasattr(m, f.name):
                setattr(m, f.name, value)


@dataclass
class GlucoseMeasurement(DeviceMeasurement):
    """Contour Care 7951H — verensokeri (mmol/L)."""
    timestamp:    Optional[datetime] = None
    glucose_mmol: Optional[float]    = None


@dataclass
class BloodPressureMeasurement(DeviceMeasurement):
    """Omron M7 — verenpaine (mmHg) ja syke."""
    timestamp: Optional[datetime] = None
    systolic:  Optional[int]      = None
    diastolic: Optional[int]      = None
    pulse:     Optional[int]      = None


@dataclass
class BodyCompositionMeasurement(DeviceMeasurement):
    """Beurer BF 720 — paino ja kehonkoostumus."""
    timestamp: Optional[datetime] = None
    weight:    Optional[float]    = None
    fat:       Optional[float]    = None
    water:     Optional[float]    = None
    muscle:    Optional[float]    = None
    bone:      Optional[float]    = None
    bmr:       Optional[int]      = None
    amr:       Optional[int]      = None


# ═══════════════════════════════════════════════════════════════════════════════
# DEVICE_TYPES — ohjelman tuntema laitelista
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass(frozen=True)
class DeviceType:
    """Yhden laitetyypin kuvaus: settings.ini-avain, näyttönimi ja mittausluokka."""
    key:             str   # settings.ini [devices]-avain, esim. "contour"
    label:           str   # ihmisluettava nimi, esim. "Contour Care 7951H"
    measurement_cls: type  # laitekohtainen mittausluokka


DEVICE_TYPES = [
    DeviceType("contour", "Contour Care 7951H", GlucoseMeasurement),
    DeviceType("omron",   "Omron M7",           BloodPressureMeasurement),
    DeviceType("beurer",  "Beurer BF 720",      BodyCompositionMeasurement),
]
DEVICE_TYPES_BY_KEY = {d.key: d for d in DEVICE_TYPES}


# ═══════════════════════════════════════════════════════════════════════════════
# MEASUREMENT — yhdistetty mittausrivi (yksi rivi Excel-taulukossa)
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class Measurement:
    """Yhden mittauskerran yhdistetyt tiedot kaikilta laitteilta.

    Yksi Measurement-olio = yksi rivi Excel-taulukossa. Laitekohtaiset
    mittausluokat yhdistetään tähän merge_into-metodilla.
    Sarakejärjestys on määritelty COLUMNS-luokkamuuttujassa.
    """
    timestamp: Optional[datetime] = None        # Contour Care 7951H
    glucose_mmol: Optional[float] = None        # Contour Care 7951H
    weight:       Optional[float] = None        # Beurer BF 720
    systolic:     Optional[int]   = None        # Omron M7
    diastolic:    Optional[int]   = None        # Omron M7
    pulse:        Optional[int]   = None        # Omron M7
    fat:          Optional[float] = None        # Beurer BF 720
    water:        Optional[float] = None        # Beurer BF 720
    muscle:       Optional[float] = None        # Beurer BF 720
    bone:         Optional[float] = None        # Beurer BF 720
    bmr:          Optional[int]   = None        # Beurer BF 720
    amr:          Optional[int]   = None        # Beurer BF 720

    # Sarakeotsikot ja leveydet Excel-taulukkoa varten
    COLUMNS = [
        ("Päiväys ja kellonaika", 22),
        ("Verensokeri (mmol/L)",  22),
        ("Paino (kg)",            12),
        ("Systolinen (mmHg)",     18),
        ("Diastolinen (mmHg)",    18),
        ("Pulssi (/min)",         14),
        ("Rasva (%)",             12),
        ("Vesi (%)",              12),
        ("Lihas (%)",             12),
        ("Luumassa (kg)",         14),
        ("BMR (kcal)",            12),
        ("AMR (kcal)",            12),
    ]

    def to_row(self) -> list:
        """Palauttaa rivin arvoina samassa järjestyksessä kuin COLUMNS."""
        ts = self.timestamp.strftime("%d.%m.%Y %H:%M") if self.timestamp else ""
        return [
            ts,
            self.glucose_mmol,
            self.weight,
            self.systolic,
            self.diastolic,
            self.pulse,
            self.fat,
            self.water,
            self.muscle,
            self.bone,
            self.bmr,
            self.amr,
        ]

    def write_row(self, ws, row: int) -> None:
        """Kirjoittaa olion tiedot Excel-taulukon riville `row`."""
        from openpyxl.styles import Alignment
        for col, value in enumerate(self.to_row(), 1):
            cell = ws.cell(row=row, column=col, value=value)
            if col > 1:
                cell.alignment = Alignment(horizontal="center")
                if isinstance(value, float):
                    cell.number_format = "0.0"


# ═══════════════════════════════════════════════════════════════════════════════
# YHDISTÄMINEN — eri laitteiden mittaukset yhteen aikaikkunaan
# ═══════════════════════════════════════════════════════════════════════════════

def combine_measurements(device_results: dict, window_minutes: int = 60) -> list:
    """Yhdistää eri laitteiden mittaukset Measurement-olioiksi.

    `device_results`: {laiteavain: [lukijan tuottamia sanakirjoja]}, esim.
        {"contour": [...], "omron": [...], "beurer": [...]}.

    Samaan aikaikkunaan (oletus 60 min) osuvat eri laitteiden mittaukset
    yhdistetään samaan Measurement-olioon. Mittaukset joille ei löydy paria
    saavat oman rivin. Tuntemattomat laiteavaimet ohitetaan.
    """
    measurements: list[Measurement] = []

    def find_or_create(ts: datetime) -> Measurement:
        if ts is None:
            m = Measurement(timestamp=None)
            measurements.append(m)
            return m
        for m in measurements:
            if m.timestamp and abs(m.timestamp - ts) <= timedelta(minutes=window_minutes):
                return m
        m = Measurement(timestamp=ts)
        measurements.append(m)
        return m

    for key, items in device_results.items():
        dtype = DEVICE_TYPES_BY_KEY.get(key)
        if dtype is None:
            continue
        for d in items:
            dm = dtype.measurement_cls.from_dict(d)
            m = find_or_create(dm.timestamp)
            dm.merge_into(m)
            if m.timestamp is None and dm.timestamp:
                m.timestamp = dm.timestamp

    measurements.sort(key=lambda x: x.timestamp or datetime.min)
    return measurements


def build_combined_measurement(device_results: dict) -> Measurement:
    """Yhdistää eri laitteiden tuoreimmat lukemat yhdeksi Measurement-riviksi.

    `device_results`: {laiteavain: [lukijan tuottamia sanakirjoja]}.

    Jokaiselta laitteelta otetaan tuorein (suurin timestamp) lukema. Rivin
    aikaleimaksi tulee tuoreimman saadun lukeman aikaleima, tai nykyhetki,
    jos yhdelläkään lukemalla ei ole aikaleimaa. Käytetään, kun kaikki laitteet
    on pollattu ja halutaan kirjoittaa yksi yhdistetty rivi taulukkoon.
    """
    m = Measurement(timestamp=datetime.now())
    latest_ts = None

    def newest(items: list) -> Optional[dict]:
        with_ts = [x for x in items if x.get("timestamp")]
        if with_ts:
            return max(with_ts, key=lambda x: x["timestamp"])
        return items[-1] if items else None

    for key, items in device_results.items():
        dtype = DEVICE_TYPES_BY_KEY.get(key)
        if dtype is None or not items:
            continue
        d = newest(items)
        if not d:
            continue
        dm = dtype.measurement_cls.from_dict(d)
        dm.merge_into(m)
        if dm.timestamp and (latest_ts is None or dm.timestamp > latest_ts):
            latest_ts = dm.timestamp

    if latest_ts is not None:
        m.timestamp = latest_ts
    return m


def combine_measurements_by_day(device_results: dict) -> list:
    """Yhdistää eri laitteiden mittaukset päivätasolla: yksi Measurement per
    kalenteripäivä.

    `device_results`: {laiteavain: [lukijan tuottamia sanakirjoja]}.

    Laite voi palauttaa monta lukemaa (esim. Contour ~20 verensokeria). Saman
    kalenteripäivän lukemat kootaan samaan Measurement-riviin, koska
    päiväpohjaisessa taulukossa on yksi rivi per päivä. Jos samalta laitteelta
    on useita saman päivän lukemia, valitaan tuorein (suurin timestamp).
    Lukemat ilman aikaleimaa ohitetaan, koska niitä ei voi sijoittaa päivän
    mukaan. Tulos järjestetään päivämäärän mukaan.
    """
    by_day: dict[date, Measurement] = {}

    for key, items in device_results.items():
        dtype = DEVICE_TYPES_BY_KEY.get(key)
        if dtype is None:
            continue

        # Valitse tältä laitteelta kunkin päivän tuorein lukema.
        newest_per_day: dict[date, dict] = {}
        for d in items:
            ts = d.get("timestamp")
            if ts is None:
                continue
            day = ts.date()
            cur = newest_per_day.get(day)
            if cur is None or ts > cur["timestamp"]:
                newest_per_day[day] = d

        for day, d in newest_per_day.items():
            m = by_day.get(day)
            if m is None:
                m = Measurement(timestamp=datetime.combine(day, datetime.min.time()))
                by_day[day] = m
            dtype.measurement_cls.from_dict(d).merge_into(m)

    result = list(by_day.values())
    result.sort(key=lambda x: x.timestamp or datetime.min)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — "Mittaukset"-välilehden kirjoitus
# ═══════════════════════════════════════════════════════════════════════════════

def save_measurements_sheet(wb, measurements: list) -> None:
    """Kirjoittaa kaikki Measurement-oliot omalle 'Mittaukset'-välilehdelle."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Mittaukset", 0)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (header, width) in enumerate(Measurement.COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="34495E")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for ri, m in enumerate(measurements, 2):
        m.write_row(ws, ri)
        for col in range(1, len(Measurement.COLUMNS) + 1):
            ws.cell(ri, col).border = border

    if not measurements:
        ws.cell(2, 1, "Ei mittauksia.")


def append_measurement(filename: str, m: Measurement) -> None:
    """Lisää yksi Measurement-rivi olemassa olevaan 'Mittaukset'-välilehteen.

    Rivi kirjoitetaan ensimmäiselle vapaalle riville otsikkorivin jälkeen.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side

    wb = load_workbook(filename)
    ws = wb["Mittaukset"]

    # Lasketaan, montako datariviä otsikon jälkeen on jo täytetty,
    # ja kirjoitetaan uusi rivi niiden jatkoksi.
    data_rows = 0
    for row in range(2, ws.max_row + 1):
        if any(ws.cell(row, col).value is not None
               for col in range(1, len(Measurement.COLUMNS) + 1)):
            data_rows += 1
    next_row = 2 + data_rows
    m.write_row(ws, next_row)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(Measurement.COLUMNS) + 1):
        ws.cell(next_row, col).border = border

    wb.save(filename)


def merge_or_append_measurements(filename: str, measurements: list) -> tuple:
    """Yhdistää uudet mittaukset olemassa olevaan 'Mittaukset'-välilehteen
    päivätasolla.

    Jokaisen Measurement-olion kohdalla:
      - Etsitään saman päivän rivit (timestamp-sarakkeesta verrataan DD.MM.YYYY-osaa).
      - Jos rivi löytyy ja kaikki sarakkeet, joihin uusi olio toisi tietoja,
        ovat tyhjiä → täytetään ne tyhjät kentät kyseiselle riville.
      - Muuten kirjoitetaan uusi rivi taulukon perään.

    Palauttaa tuplen (päivitettyjä rivejä, uusia rivejä).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side

    wb = load_workbook(filename)
    ws = wb["Mittaukset"]
    n_cols = len(Measurement.COLUMNS)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Siivoa pois mahdollinen "Ei mittauksia." -placeholder
    if ws.cell(2, 1).value == "Ei mittauksia.":
        ws.cell(2, 1).value = None

    def next_empty_row() -> int:
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            if any(ws.cell(row, col).value is not None
                   for col in range(1, n_cols + 1)):
                data_rows += 1
        return 2 + data_rows

    def add_borders(row: int) -> None:
        for col in range(1, n_cols + 1):
            ws.cell(row, col).border = border

    updated, added = 0, 0

    for m in measurements:
        new_values = m.to_row()
        target_date = (m.timestamp.strftime("%d.%m.%Y")
                       if m.timestamp else None)
        merged = False

        if target_date is not None:
            for row in range(2, ws.max_row + 1):
                ts_str = ws.cell(row, 1).value
                if not (isinstance(ts_str, str) and ts_str.startswith(target_date)):
                    continue

                # Ristiriita: uusi olio yrittää kirjoittaa sarakkeeseen, jossa
                # on jo arvo. Aikaleimasarake (1) ohitetaan, sitä ei päivitetä.
                conflict = any(
                    new_values[col - 1] not in (None, "")
                    and ws.cell(row, col).value is not None
                    for col in range(2, n_cols + 1)
                )
                if conflict:
                    continue

                # Täytä tyhjät kentät
                for col in range(2, n_cols + 1):
                    val = new_values[col - 1]
                    if val in (None, ""):
                        continue
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.alignment = Alignment(horizontal="center")
                    if isinstance(val, float):
                        cell.number_format = "0.0"
                    cell.border = border
                merged = True
                updated += 1
                break

        if not merged:
            row = next_empty_row()
            m.write_row(ws, row)
            add_borders(row)
            added += 1

    wb.save(filename)
    return updated, added


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — kirjoitus päiväpohjaiseen taulukkoon (yksi rivi per päivä)
# ═══════════════════════════════════════════════════════════════════════════════

def _cell_to_date(value) -> Optional[date]:
    """Tulkitsee ensimmäisen sarakkeen solun kalenteripäiväksi.

    Hyväksyy datetime/date-oliot ja yleisimmät merkkijonomuodot (esim.
    'YYYY-MM-DD' tai 'DD.MM.YYYY'). Palauttaa None, jos solua ei voi tulkita
    päivämääräksi (esim. otsikkorivi)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                    "%d.%m.%Y %H:%M", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def find_measurement_sheet(wb):
    """Etsii työkirjasta välilehden, jonka ensimmäisen sarakkeen otsikko on
    päiväyssarake (alkaa 'Päiväys'). Palauttaa aktiivisen välilehden, jos
    otsikkoa ei löydy mistään (esim. eri kielinen taulukko)."""
    for ws in wb.worksheets:
        v = ws.cell(1, 1).value
        if isinstance(v, str) and v.strip().startswith("Päiväys"):
            return ws
    return wb.active


def write_measurements_by_date(filename: str, measurements: list) -> tuple:
    """Kirjoittaa mittaukset taulukkoon, jonka ensimmäisessä sarakkeessa on
    valmiiksi jokaisen päivän päiväys (vuositiedostossa 1.1.–31.12., kuukausi-
    tiedostossa kuun ensimmäisestä viimeiseen päivään).

    Jokaiselle mittaukselle etsitään riviltä sama kalenteripäivä ja siihen
    kirjoitetaan mittauksen ei-tyhjät arvot (sarakkeet 2…N). Päiväyssaraketta
    (1) ei muuteta. Sarakkeet, joille mittauksella ei ole arvoa, jätetään
    ennalleen — esim. pelkkä verensokerimittaus ei pyyhi saman päivän painoa.
    Solu täytetään vain, jos se on tyhjä: taulukossa jo oleva arvo säilyy
    aina ennallaan, eikä uusi lukema ylikirjoita sitä.

    Palauttaa tuplen (kirjoitettuja rivejä, sijoittamatta jääneet mittaukset),
    jossa jälkimmäinen on lista mittauksista, joiden päivää ei löytynyt
    taulukosta (esim. väärä vuosi tai kuukausi)."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side

    wb = load_workbook(filename)
    ws = find_measurement_sheet(wb)
    n_cols = len(Measurement.COLUMNS)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Indeksoi taulukon rivit päivämäärän mukaan (ensimmäinen osuma per päivä).
    date_to_row: dict[date, int] = {}
    for row in range(2, ws.max_row + 1):
        d = _cell_to_date(ws.cell(row, 1).value)
        if d is not None and d not in date_to_row:
            date_to_row[d] = row

    written = 0
    unmatched = []

    for m in measurements:
        day = m.timestamp.date() if m.timestamp else None
        row = date_to_row.get(day) if day is not None else None
        if row is None:
            unmatched.append(m)
            continue

        values = m.to_row()          # [0] = aikaleima, [1…] = mittausarvot
        for col in range(2, n_cols + 1):
            val = values[col - 1]
            if val in (None, ""):
                continue
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):    # solussa jo arvo — ei ylikirjoiteta
                continue
            cell.value = val
            cell.alignment = Alignment(horizontal="center")
            if isinstance(val, float):
                cell.number_format = "0.0"
            cell.border = border
        written += 1

    wb.save(filename)
    return written, unmatched