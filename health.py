#!/usr/bin/env python3
"""
Terveysmittaukset BLE-lukija — Linux
=====================================
Tukee:
  • Ascensia Contour Care 7951H       — verensokeri (mmol/L)
  • Omron M7 Intelli IT (HEM-7361T)   — verenpaine (mmHg) + syke
  • Beurer BF 720                      — paino, rasva%, vesi%, lihas%,
                                         luumassa, BMR, AMR

Protokollat:
  • Contour: Bluetooth GATT Glucose Service (UUID 0x1808)
  • Omron:   omblepy-projektin BLE-protokolla
  • Beurer:  openScale-projektin reverse-engineered BF105/BF720-protokolla
             (service 0xFFE0, characteristic 0xFFE1)

Asennus:
    pip install bleak openpyxl

Käyttö:
    # Etsi laitteet
    python3 health.py --scan

    # Lue verensokeri
    python3 health.py --contour AA:BB:CC:DD:EE:FF

    # Lue verenpaine
    python3 health.py --omron AA:BB:CC:DD:EE:FF

    # Lue paino + kehonkoostumus
    python3 health.py --beurer AA:BB:CC:DD:EE:FF

    # Lue kaikki kolme samaan tiedostoon
    python3 health.py \\
        --contour AA:BB:CC:DD:EE:FF \\
        --omron   BB:CC:DD:EE:FF:GG \\
        --beurer  CC:DD:EE:FF:GG:HH

    # Oma tiedostonimi
    python3 health.py --beurer AA:BB:CC:DD:EE:FF --excel paino_2026.xlsx

Paritusohjeet:
  Contour:  bluetoothctl → pair XX:XX → trust XX:XX  (mittari näyttää passkeyn)
  Omron M7: pidä BT-nappia pohjassa kunnes P- vilkkuu → aja --omron
  Beurer:   pidä ON-nappia 3 sek kunnes BT-symboli vilkkuu → aja --beurer
            (tai astu vaakaan — se aktivoi BT:n automaattisesti)
"""

import asyncio, struct, argparse, sys, os
from datetime import datetime

try:
    from bleak import BleakScanner, BleakClient
    from bleak.exc import BleakError, BleakDeviceNotFoundError
except ImportError:
    print("Virhe: bleak-kirjasto puuttuu. Asenna: pip install bleak")
    sys.exit(1)

from measurement import (
    Measurement,
    DEVICE_TYPES,
    combine_measurements,
    combine_measurements_by_day,
    save_measurements_sheet,
    write_measurements_by_date,
)

# ── Contour GATT UUID:t ──────────────────────────────────────────────────────
GLUCOSE_MEASUREMENT_UUID = "00002a18-0000-1000-8000-00805f9b34fb"
GLUCOSE_CONTEXT_UUID     = "00002a34-0000-1000-8000-00805f9b34fb"
RACP_UUID                = "00002a52-0000-1000-8000-00805f9b34fb"
RACP_REPORT_ALL          = bytes([0x01, 0x01])

# ── Omron M7 BLE UUID:t (standardi GATT Blood Pressure Service 0x1810) ───────
# Löydetty bluetoothctl list-attributes -komennolla HEM-7361T:stä
# Blood Pressure Measurement (notify) + Record Access Control Point (write+notify)
OMRON_BP_MEASUREMENT_UUID = "00002a35-0000-1000-8000-00805f9b34fb"
OMRON_RACP_UUID           = "00002a52-0000-1000-8000-00805f9b34fb"
OMRON_RACP_REPORT_ALL     = bytes([0x01, 0x01])  # Report all stored records

# ── Beurer BF 720 BLE UUID:t (openScale BF105/BF720-protokolla) ─────────────
# Protokolla reverse-engineered openScale-projektissa:
# https://github.com/oliexdev/openScale
# Service 0xFFE0, Characteristic 0xFFE1
# Pakettiformaatti (20 tavua):
#   [0]     tunniste: 0x10=reaaliaikainen, 0x1F/0x1E=historiatietue
#   [1]     vuosi - 2000
#   [2]     kuukausi
#   [3]     päivä
#   [4]     tunti
#   [5]     minuutti
#   [6]     sekunti
#   [7]     käyttäjänumero (1-8)
#   [8-9]   paino * 10 (uint16 LE, esim. 756 = 75.6 kg)
#   [10-11] rasvaprosentti * 10 (uint16 LE)
#   [12-13] vesiprosentti * 10 (uint16 LE)
#   [14-15] lihasprosentti * 10 (uint16 LE)
#   [16]    luumassa * 10 (uint8)
#   [17-18] BMR kcal (uint16 LE)
#   [19]    AMR-aktiivisuustaso (1-5)
BEURER_CHAR_UUID         = "0000ffe1-0000-1000-8000-00805f9b34fb"
BEURER_CMD_GET_DATA      = bytes([0x21, 0x01])

# PAL-kertoimet AMR-laskentaan
PAL = {1: 1.2, 2: 1.375, 3: 1.55, 4: 1.725, 5: 1.9}

# ── Globaalit tulospuskurit ───────────────────────────────────────────────────
glucose_data  = []
bp_data       = []
scale_data    = []
racp_done     = asyncio.Event()
omron_done    = asyncio.Event()
beurer_done   = asyncio.Event()
omron_replies = []  # ei enää käytetty, pidetään yhteensopivuuden vuoksi


# ═══════════════════════════════════════════════════════════════════════════════
# CONTOUR — verensokeri
# ═══════════════════════════════════════════════════════════════════════════════

def parse_glucose(data: bytes) -> dict | None:
    if len(data) < 10:
        return None
    flags = data[0]
    seq   = struct.unpack_from("<H", data, 1)[0]
    try:
        year, month, day, hour, minute, second = struct.unpack_from("<HBBBBB", data, 3)
        timestamp = datetime(year, month, day, hour, minute, second)
    except (ValueError, struct.error):
        timestamp = None
    offset = 10
    if flags & 0x01 and len(data) >= offset + 2:
        offset += 2
    glucose_mmol = None
    if (flags & 0x02) and len(data) >= offset + 2:
        raw      = struct.unpack_from("<H", data, offset)[0]
        mantissa = raw & 0x0FFF
        if mantissa & 0x0800: mantissa -= 0x1000
        exponent = (raw >> 12) & 0x0F
        if exponent & 0x08:   exponent -= 0x10
        value = mantissa * (10 ** exponent)
        glucose_mmol = value * 1000
        if glucose_mmol and glucose_mmol > 40:
            glucose_mmol = value * 1000 * 0.0555
    if not glucose_mmol:
        return None
    return {"seq": seq, "timestamp": timestamp,
            "glucose_mmol": round(glucose_mmol, 1)}


def on_glucose(sender, data: bytearray):
    result = parse_glucose(bytes(data))
    if result:
        glucose_data.append(result)
        ts = result["timestamp"].strftime("%d.%m.%Y %H:%M") if result["timestamp"] else "?"
        print(f"  Verensokeri [{result['seq']:3d}] {ts}  →  {result['glucose_mmol']:.1f} mmol/L")


def on_racp(sender, data: bytearray):
    if len(data) >= 1 and data[0] == 0x06:
        racp_done.set()


async def read_contour(mac: str) -> list:
    global glucose_data
    glucose_data = []
    racp_done.clear()
    print(f"\n[Contour] Yhdistetään {mac} ...")
    print("[Contour] Aktivoi mittarin BT: OK pitkään → ylänuoli 3 sek → )) palaa\n")
    try:
        async with BleakClient(mac, timeout=20.0) as client:
            if not client.is_connected:
                print("[Contour] Yhteys epäonnistui."); return []
            print("[Contour] Yhdistetty! Luetaan verensokeriarvoja...\n")
            await client.start_notify(GLUCOSE_MEASUREMENT_UUID, on_glucose)
            await client.start_notify(RACP_UUID, on_racp)
            try:
                await client.start_notify(GLUCOSE_CONTEXT_UUID, lambda s, d: None)
            except Exception:
                pass
            await client.write_gatt_char(RACP_UUID, RACP_REPORT_ALL, response=True)
            try:
                await asyncio.wait_for(racp_done.wait(), timeout=30.0)
            except asyncio.TimeoutError:
                print("[Contour] Aikakatkaisu — käytetään saadut arvot.")
            await client.stop_notify(GLUCOSE_MEASUREMENT_UUID)
            await client.stop_notify(RACP_UUID)
    except BleakDeviceNotFoundError:
        print("[Contour] Mittaria ei löydy. Onko BT päällä mittarissa?")
    except BleakError as e:
        print(f"[Contour] BLE-virhe: {e}")
    print(f"[Contour] Saatiin {len(glucose_data)} verensokeriarvoa.")
    return glucose_data


# ═══════════════════════════════════════════════════════════════════════════════
# OMRON M7 — verenpaine ja syke
# ═══════════════════════════════════════════════════════════════════════════════

def parse_omron_bp(data: bytes) -> dict | None:
    """
    Purkaa GATT 2A35 Blood Pressure Measurement -paketin (IEEE 11073 SFLOAT).
    Paketti:
      [0]      flags
      [1-2]    systolinen SFLOAT (LE)
      [3-4]    diastolinen SFLOAT (LE)
      [5-6]    MAP SFLOAT (LE)
      [7-13]   aikaleima (vuosi uint16, kk, pv, h, min, s)  — jos lippu 1 asetettu
      [14-15]  syke SFLOAT (LE)                             — jos lippu 2 asetettu
      [16]     käyttäjätunnus                               — jos lippu 3 asetettu
    """
    if len(data) < 7:
        return None

    flags = data[0]

    def sfloat(offset):
        """Purkaa IEEE 11073 SFLOAT (2 tavua LE) → float."""
        raw = struct.unpack_from("<H", data, offset)[0]
        mantissa = raw & 0x0FFF
        if mantissa & 0x0800:
            mantissa -= 0x1000
        exponent = (raw >> 12) & 0x0F
        if exponent & 0x08:
            exponent -= 0x10
        return mantissa * (10 ** exponent)

    # Yksikkö: lippu 0 = mmHg (0), kPa (1)
    unit_kpa = bool(flags & 0x01)

    systolic  = sfloat(1)
    diastolic = sfloat(3)
    # MAP (mean arterial pressure) offset 5 — ei tallenneta

    offset = 7

    # Aikaleima — lippu 1
    timestamp = None
    if flags & 0x02:
        if len(data) >= offset + 7:
            year   = struct.unpack_from("<H", data, offset)[0]
            month  = data[offset+2]
            day    = data[offset+3]
            hour   = data[offset+4]
            minute = data[offset+5]
            second = data[offset+6]
            try:
                timestamp = datetime(year, month, day, hour, minute, second)
            except ValueError:
                timestamp = None
        offset += 7

    # Syke — lippu 2
    pulse = None
    if flags & 0x04:
        if len(data) >= offset + 2:
            pulse = int(sfloat(offset))
        offset += 2

    # Käyttäjätunnus — lippu 3
    user = 1
    if flags & 0x08:
        if len(data) > offset:
            user = data[offset]
        offset += 1

    # Muunna kPa → mmHg jos tarpeellista
    if unit_kpa:
        systolic  *= 7.50062
        diastolic *= 7.50062

    systolic  = round(systolic)
    diastolic = round(diastolic)

    # Saniteettitarkistus
    if not (50 <= systolic <= 300 and 30 <= diastolic <= 200):
        return None
    if pulse and not (30 <= pulse <= 220):
        pulse = None

    return {
        "seq":       len(bp_data),  # järjestysnumero
        "timestamp": timestamp,
        "systolic":  systolic,
        "diastolic": diastolic,
        "pulse":     pulse or None,
        "user":      user,
    }


def on_omron_bp_notify(sender, data: bytearray):
    """Notifikaatio Blood Pressure Measurement -characteristicista."""
    result = parse_omron_bp(bytes(data))
    if result:
        bp_data.append(result)
        ts = result["timestamp"].strftime("%d.%m.%Y %H:%M") if result["timestamp"] else "?"
        print(f"  Verenpaine [{result['seq']:3d}] {ts}  "
              f"→  {result['systolic']}/{result['diastolic']} mmHg  "
              f"syke {result['pulse'] or '-'} /min")


def on_omron_racp_notify(sender, data: bytearray):
    """RACP-vastaus — 0x06 = kaikki tietueet lähetetty."""
    raw = bytes(data)
    if len(raw) >= 1 and raw[0] == 0x06:
        omron_done.set()


async def read_omron(mac: str) -> list:
    """Yhdistä Omron M7 -mittariin ja lue verenpaine + syke GATT BP Service -protokollalla."""
    global bp_data
    bp_data = []
    omron_done.clear()
    print(f"\n[Omron] Yhdistetään {mac} ...")
    print("[Omron] Aktivoi mittarin BT: pidä BT-nappia pohjassa kunnes P- vilkkuu.\n")
    try:
        async with BleakClient(mac, timeout=20.0) as client:
            if not client.is_connected:
                print("[Omron] Yhteys epäonnistui."); return []
            print("[Omron] Yhdistetty! Luetaan verenpainearvoja...\n")

            # Tilaa Blood Pressure Measurement -notifikaatiot
            await client.start_notify(OMRON_BP_MEASUREMENT_UUID, on_omron_bp_notify)
            # Tilaa RACP-notifikaatiot (siirron valmistuminen)
            await client.start_notify(OMRON_RACP_UUID, on_omron_racp_notify)

            # Pyydä kaikki tallennetut mittaukset
            await client.write_gatt_char(OMRON_RACP_UUID,
                                         OMRON_RACP_REPORT_ALL, response=True)

            try:
                await asyncio.wait_for(omron_done.wait(), timeout=30.0)
            except asyncio.TimeoutError:
                print("[Omron] Aikakatkaisu — käytetään saadut arvot.")

            await client.stop_notify(OMRON_BP_MEASUREMENT_UUID)
            await client.stop_notify(OMRON_RACP_UUID)

    except BleakDeviceNotFoundError:
        print("[Omron] Mittaria ei löydy. Onko BT päällä?")
    except BleakError as e:
        print(f"[Omron] BLE-virhe: {e}")
    print(f"[Omron] Saatiin {len(bp_data)} verenpainearvoa.")
    return bp_data


# ═══════════════════════════════════════════════════════════════════════════════
# BEURER BF 720 — paino ja kehonkoostumus
# ═══════════════════════════════════════════════════════════════════════════════

def parse_beurer_record(data: bytes) -> dict | None:
    """Purkaa Beurer BF 720 BLE-paketin (openScale BF105-protokolla)."""
    if len(data) < 17:
        return None
    if data[0] not in (0x10, 0x1E, 0x1F):
        return None
    try:
        timestamp = datetime(2000 + data[1], data[2], data[3],
                             data[4], data[5],
                             data[6] if len(data) > 6 else 0)
    except (ValueError, IndexError):
        timestamp = None

    def u16(offset):
        return struct.unpack_from("<H", data, offset)[0] if len(data) >= offset + 2 else None
    def u8(offset):
        return data[offset] if len(data) > offset else None

    weight_raw = u16(8);  fat_raw = u16(10); water_raw = u16(12)
    muscle_raw = u16(14); bone_raw = u8(16); bmr_raw   = u16(17)
    amr_level  = u8(19)

    weight = round(weight_raw / 10, 1) if weight_raw else None
    if not weight or not (20.0 <= weight <= 300.0):
        return None

    fat    = round(fat_raw    / 10, 1) if fat_raw    else None
    water  = round(water_raw  / 10, 1) if water_raw  else None
    muscle = round(muscle_raw / 10, 1) if muscle_raw else None
    bone   = round(bone_raw   / 10, 1) if bone_raw   else None
    bmr    = bmr_raw if bmr_raw else None
    amr    = round(bmr * PAL.get(amr_level, 1.2)) if bmr and amr_level else None

    return {
        "timestamp": timestamp,
        "user":      u8(7) or 1,
        "weight":    weight,
        "fat":       fat,
        "water":     water,
        "muscle":    muscle,
        "bone":      bone,
        "bmr":       bmr,
        "amr":       amr,
    }


def on_beurer_notify(sender, data: bytearray):
    raw = bytes(data)
    result = parse_beurer_record(raw)
    if result:
        # Vältä duplikaatit
        if not any(m["timestamp"] == result["timestamp"] and
                   m["weight"] == result["weight"] for m in scale_data):
            scale_data.append(result)
            ts = result["timestamp"].strftime("%d.%m.%Y %H:%M") if result["timestamp"] else "?"
            fat_s    = f"  rasva {result['fat']}%"    if result["fat"]    else ""
            water_s  = f"  vesi {result['water']}%"   if result["water"]  else ""
            muscle_s = f"  lihas {result['muscle']}%" if result["muscle"] else ""
            print(f"  Vaaka {ts}  →  {result['weight']} kg{fat_s}{water_s}{muscle_s}")
    if len(raw) >= 1 and raw[0] == 0xFF:
        beurer_done.set()


async def read_beurer(mac: str) -> list:
    """Yhdistä Beurer BF 720 -vaakaan ja lue kehonkoostumusmittaukset."""
    global scale_data
    scale_data = []
    beurer_done.clear()
    print(f"\n[Beurer] Odotetaan vaakaa {mac} ...")
    print("[Beurer] Astu vaakalle — ohjelma yhdistää automaattisesti kun vaaka herää.\n")

    # Odotetaan että vaaka ilmestyy skannaukseen (max 60 sekuntia)
    print("[Beurer] Skannataan", end="", flush=True)
    found = False
    for _ in range(30):  # 30 x 2 sek = 60 sek
        devices = await BleakScanner.discover(timeout=2.0)
        for d in devices:
            if d.address.upper() == mac.upper():
                found = True
                break
        if found:
            break
        print(".", end="", flush=True)
    print()

    if not found:
        print("[Beurer] Vaakaa ei löydy 60 sekunnin aikana. Astu vaakalle ja yritä uudelleen.")
        return []

    print("[Beurer] Vaaka löytyi! Yhdistetään...\n")
    try:
        async with BleakClient(mac, timeout=20.0) as client:
            if not client.is_connected:
                print("[Beurer] Yhteys epäonnistui."); return []
            print("[Beurer] Yhdistetty! Odotetaan mittauksia...\n")
            await client.start_notify(BEURER_CHAR_UUID, on_beurer_notify)
            try:
                await client.write_gatt_char(BEURER_CHAR_UUID,
                                             BEURER_CMD_GET_DATA, response=False)
            except Exception:
                pass
            try:
                await asyncio.wait_for(beurer_done.wait(), timeout=40.0)
            except asyncio.TimeoutError:
                print("[Beurer] Aikakatkaisu — käytetään saadut arvot.")
            await client.stop_notify(BEURER_CHAR_UUID)
    except BleakDeviceNotFoundError:
        print("[Beurer] Vaakaa ei löydy. Onko BT päällä vaakassa?")
    except BleakError as e:
        print(f"[Beurer] BLE-virhe: {e}")
    print(f"[Beurer] Saatiin {len(scale_data)} mittausta.")
    return scale_data


# ═══════════════════════════════════════════════════════════════════════════════
# SKANNAUS
# ═══════════════════════════════════════════════════════════════════════════════

async def scan_devices():
    print("Etsitään Bluetooth-laitteita 10 sekuntia...")
    print("Aktivoi mittarien/vaa'an BT nyt!\n")

    # Uudempi bleak palauttaa (device, advertisement_data) -pareja
    results = []
    def cb(device, adv_data):
        results.append((device, adv_data))
    async with BleakScanner(cb) as scanner:
        await asyncio.sleep(10.0)

    keywords = ["CONTOUR", "BAYER", "ASCENSIA", "OMRON", "HEM-",
                "BGM", "BLESMART", "BF105", "BF720", "BF700", "BEURER"]

    # Deduplikoi MAC-osoitteen perusteella
    seen = {}
    for d, adv in results:
        seen[d.address] = (d, adv)

    known = [(d, adv) for d, adv in seen.values()
             if any(k in (d.name or "").upper() for k in keywords)]

    if known:
        print("=== Tunnistetut terveysmittarit ===")
        for d, adv in known:
            print(f"  {d.address}  {d.name or 'tuntematon'}")

    print("\n=== Kaikki lähistön laitteet (vahvimmat ensin) ===")
    for d, adv in sorted(seen.values(),
                         key=lambda x: x[1].rssi if x[1].rssi else -999,
                         reverse=True)[:20]:
        rssi = adv.rssi if adv.rssi else "?"
        print(f"  {d.address}  RSSI:{rssi!s:4}  {d.name or ''}")
    print("\nKäytä --contour, --omron tai --beurer MAC-osoitteen kanssa.")


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL / LIBREOFFICE TALLENNUS
# ═══════════════════════════════════════════════════════════════════════════════

def save_to_excel(glucose: list, bp: list, scale: list, filename: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Virhe: openpyxl puuttuu. Asenna: pip install openpyxl")
        return

    wb = Workbook()
    # Poista oletusvälilehti — luomme välilehdet itse
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ── Mittaukset-välilehti (yhdistetty taulukko, Measurement-oliot) ────────
    measurements = combine_measurements({
        "contour": glucose,
        "omron":   bp,
        "beurer":  scale,
    })
    save_measurements_sheet(wb, measurements)

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, value, bg="2E75B6"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        return c

    def line_chart(ws, cols, title, y_label, anchor, n, colors):
        chart = LineChart()
        chart.title = title; chart.style = 10
        chart.y_axis.title = y_label
        chart.height = 12;   chart.width  = 24
        for col, color in zip(cols, colors):
            ref = Reference(ws, min_col=col, min_row=1, max_row=n + 1)
            chart.add_data(ref, titles_from_data=True)
        for i, color in enumerate(colors):
            chart.series[i].graphicalProperties.line.solidFill = color
        ws.add_chart(chart, anchor)

    # ── 1. Verensokeri ────────────────────────────────────────────────────────
    ws_g = wb.create_sheet("Verensokeri")
    if glucose:
        for col, (h, w) in enumerate(zip(
            ["Järj.nro","Päivämäärä","Kellonaika","Verensokeri (mmol/L)","Huomio"],
            [10, 15, 12, 22, 18]), 1):
            hdr(ws_g, 1, col, h)
            ws_g.column_dimensions[get_column_letter(col)].width = w
        ws_g.row_dimensions[1].height = 22
        lf = PatternFill("solid", start_color="FFE0E0")
        hf = PatternFill("solid", start_color="FFEAC0")
        of = PatternFill("solid", start_color="E8F5E9")
        for ri, m in enumerate(sorted(glucose, key=lambda x: x["timestamp"] or datetime.min), 2):
            ts = m["timestamp"]; glc = m["glucose_mmol"]
            ws_g.cell(ri,1, m["seq"])
            ws_g.cell(ri,2, ts.strftime("%d.%m.%Y") if ts else "")
            ws_g.cell(ri,3, ts.strftime("%H:%M")    if ts else "")
            gc = ws_g.cell(ri,4, glc)
            gc.number_format = "0.0"; gc.alignment = Alignment(horizontal="center")
            if glc is None:    fill, note = None, ""
            elif glc < 4.0:    fill, note = lf,   "⚠ Matala"
            elif glc > 8.0:    fill, note = hf,   "▲ Korkea"
            else:              fill, note = of,   "✓ OK"
            ws_g.cell(ri,5, note)
            for col in range(1,6):
                c = ws_g.cell(ri,col); c.border = border
                if fill: c.fill = fill
        line_chart(ws_g, [4], "Verensokeri aikajärjestyksessä",
                   "mmol/L", "G2", len(glucose), ["2E75B6"])
    else:
        ws_g.cell(1,1, "Ei verensokeriarvoja.")

    # ── 2. Verenpaine ─────────────────────────────────────────────────────────
    ws_b = wb.create_sheet("Verenpaine")
    if bp:
        for col, (h, w) in enumerate(zip(
            ["Järj.nro","Päivämäärä","Kellonaika",
             "Systolinen (mmHg)","Diastolinen (mmHg)","Syke (/min)","Käyttäjä","Huomio"],
            [10,15,12,18,18,14,10,18]), 1):
            hdr(ws_b,1,col,h,bg="C0392B")
            ws_b.column_dimensions[get_column_letter(col)].width = w
        ws_b.row_dimensions[1].height = 22
        lf = PatternFill("solid", start_color="FFE0E0")
        hf = PatternFill("solid", start_color="FFEAC0")
        of = PatternFill("solid", start_color="E8F5E9")
        for ri, m in enumerate(sorted(bp, key=lambda x: x["timestamp"] or datetime.min), 2):
            ts = m["timestamp"]
            ws_b.cell(ri,1, m["seq"])
            ws_b.cell(ri,2, ts.strftime("%d.%m.%Y") if ts else "")
            ws_b.cell(ri,3, ts.strftime("%H:%M")    if ts else "")
            for col, val in zip([4,5,6,7],[m["systolic"],m["diastolic"],m["pulse"],m["user"]]):
                ws_b.cell(ri,col,val).alignment = Alignment(horizontal="center")
            if m["systolic"] >= 140 or m["diastolic"] >= 90: fill, note = hf, "▲ Korkea"
            elif m["systolic"] < 90 or m["diastolic"] < 60:  fill, note = lf, "⚠ Matala"
            else:                                              fill, note = of, "✓ Normaali"
            ws_b.cell(ri,8, note)
            for col in range(1,9):
                c = ws_b.cell(ri,col); c.border = border; c.fill = fill
        line_chart(ws_b, [4,5], "Verenpaine aikajärjestyksessä",
                   "mmHg", "J2", len(bp), ["C0392B","E74C3C"])
    else:
        ws_b.cell(1,1, "Ei verenpainearvoja.")

    # ── 3. Paino ja kehonkoostumus (Beurer BF 720) ───────────────────────────
    ws_s = wb.create_sheet("Paino")
    if scale:
        for col, (h, w) in enumerate(zip(
            ["Päivämäärä","Kellonaika","Paino (kg)","Rasva (%)","Vesi (%)",
             "Lihas (%)","Luumassa (kg)","BMR (kcal)","AMR (kcal)","Käyttäjä"],
            [15,12,12,12,12,12,14,12,12,10]), 1):
            hdr(ws_s,1,col,h,bg="27AE60")
            ws_s.column_dimensions[get_column_letter(col)].width = w
        ws_s.row_dimensions[1].height = 22
        ok_f = PatternFill("solid", start_color="E8F5E9")
        for ri, m in enumerate(sorted(scale, key=lambda x: x["timestamp"] or datetime.min), 2):
            ts = m["timestamp"]
            ws_s.cell(ri,1, ts.strftime("%d.%m.%Y") if ts else "")
            ws_s.cell(ri,2, ts.strftime("%H:%M")    if ts else "")
            for col, key in enumerate(
                    ["weight","fat","water","muscle","bone","bmr","amr","user"], 3):
                val = m.get(key)
                c = ws_s.cell(ri,col,val)
                c.alignment = Alignment(horizontal="center")
                if isinstance(val, float): c.number_format = "0.0"
            for col in range(1,11):
                c = ws_s.cell(ri,col); c.border = border; c.fill = ok_f
        line_chart(ws_s, [3], "Paino aikajärjestyksessä",
                   "kg", "L2", len(scale), ["27AE60"])
    else:
        ws_s.cell(1,1, "Ei painomittauksia — lisää --beurer MAC-osoite kun vaaka on hankittu.")

    # ── 4. Yhteenveto ─────────────────────────────────────────────────────────
    ws_y = wb.create_sheet("Yhteenveto")
    ws_y["A1"] = "Yhteenveto — Terveysmittaukset"
    ws_y["A1"].font = Font(bold=True, size=14)
    ws_y.column_dimensions["A"].width = 30
    ws_y.column_dimensions["B"].width = 15

    def section(ws, start, title, rows, bg):
        c = ws.cell(start, 1, title)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=bg)
        ws.merge_cells(f"A{start}:B{start}")
        for r, (label, formula) in enumerate(rows, start+1):
            ws.cell(r, 1, label).font = Font(bold=True)
            ws.cell(r, 2, formula)
        return start + len(rows) + 2

    ge = len(glucose)+1 if glucose else 2
    be = len(bp)+1      if bp      else 2
    se = len(scale)+1   if scale   else 2

    nxt = section(ws_y, 2, "Verensokeri (Contour Care)", [
        ("Mittauksia",         f"=COUNTA(Verensokeri!D2:D{ge})"),
        ("Keskiarvo (mmol/L)", f"=IFERROR(ROUND(AVERAGE(Verensokeri!D2:D{ge}),1),\"-\")"),
        ("Minimi (mmol/L)",    f"=IFERROR(MIN(Verensokeri!D2:D{ge}),\"-\")"),
        ("Maksimi (mmol/L)",   f"=IFERROR(MAX(Verensokeri!D2:D{ge}),\"-\")"),
        ("Matalia (<4.0)",     f'=IFERROR(COUNTIF(Verensokeri!D2:D{ge},"<4"),0)'),
        ("Korkeita (>8.0)",    f'=IFERROR(COUNTIF(Verensokeri!D2:D{ge},">8"),0)'),
    ], "2E75B6")

    nxt = section(ws_y, nxt, "Verenpaine (Omron M7)", [
        ("Mittauksia",          f"=COUNTA(Verenpaine!D2:D{be})"),
        ("Keskim. systolinen",  f"=IFERROR(ROUND(AVERAGE(Verenpaine!D2:D{be}),0),\"-\")"),
        ("Keskim. diastolinen", f"=IFERROR(ROUND(AVERAGE(Verenpaine!E2:E{be}),0),\"-\")"),
        ("Keskim. syke",        f"=IFERROR(ROUND(AVERAGE(Verenpaine!F2:F{be}),0),\"-\")"),
        ("Korkea (≥140/90)",    f'=IFERROR(COUNTIFS(Verenpaine!D2:D{be},">=140"),0)'),
        ("Normaali (<140/90)",  f'=IFERROR(COUNTIFS(Verenpaine!D2:D{be},"<140",Verenpaine!E2:E{be},"<90"),0)'),
    ], "C0392B")

    section(ws_y, nxt, "Paino ja kehonkoostumus (Beurer BF 720)", [
        ("Mittauksia",          f"=COUNTA(Paino!C2:C{se})"),
        ("Keskim. paino (kg)",  f"=IFERROR(ROUND(AVERAGE(Paino!C2:C{se}),1),\"-\")"),
        ("Minimi paino (kg)",   f"=IFERROR(MIN(Paino!C2:C{se}),\"-\")"),
        ("Maksimi paino (kg)",  f"=IFERROR(MAX(Paino!C2:C{se}),\"-\")"),
        ("Keskim. rasva (%)",   f"=IFERROR(ROUND(AVERAGE(Paino!D2:D{se}),1),\"-\")"),
        ("Keskim. vesi (%)",    f"=IFERROR(ROUND(AVERAGE(Paino!E2:E{se}),1),\"-\")"),
        ("Keskim. BMR (kcal)",  f"=IFERROR(ROUND(AVERAGE(Paino!H2:H{se}),0),\"-\")"),
    ], "27AE60")

    wb.save(filename)
    print(f"\nTiedosto tallennettu: {filename}")
    print(f"Avaa: libreoffice --calc {filename}")


def create_new_file(filename: str, period: str = "year",
                    ref_date: datetime | None = None) -> None:
    """Luo uuden Excel-tiedoston 'Mittaukset'-välilehdellä, jonka ensimmäiseen
    sarakkeeseen on valmiiksi kirjoitettu jokaisen päivän päiväys. Mittaukset
    sijoitetaan myöhemmin oikealle päiväriville (write_measurements_by_date).

    period="year"  → koko vuoden päivät (1.1.–31.12.)
    period="month" → kyseisen kuukauden päivät (1. – kuun viimeinen päivä)

    Päivät luodaan ref_date-vuoden (ja -kuukauden) mukaan; oletuksena nykyhetki.
    Tarvittavat hakemistot luodaan automaattisesti.
    """
    from openpyxl import Workbook
    from datetime import date, timedelta

    target_dir = os.path.dirname(filename)
    if target_dir:
        os.makedirs(target_dir, exist_ok=True)

    ref = ref_date or datetime.now()
    if period == "month":
        start = date(ref.year, ref.month, 1)
        if ref.month == 12:
            end = date(ref.year, 12, 31)
        else:
            end = date(ref.year, ref.month + 1, 1) - timedelta(days=1)
    else:
        start = date(ref.year, 1, 1)
        end = date(ref.year, 12, 31)

    wb = Workbook()
    wb.remove(wb.active)
    save_measurements_sheet(wb, [])  # luo välilehden otsikkorivillä
    # save_measurements_sheet kirjoittaa tyhjään tauluun "Ei mittauksia." -solun;
    # poistetaan se ennen päivärivien kirjoittamista.
    ws = wb["Mittaukset"]
    if ws.cell(2, 1).value == "Ei mittauksia.":
        ws.cell(2, 1).value = None

    # Yksi rivi per päivä; päiväys ensimmäiseen sarakkeeseen.
    row = 2
    d = start
    while d <= end:
        cell = ws.cell(row=row, column=1, value=datetime(d.year, d.month, d.day))
        cell.number_format = "YYYY-MM-DD"
        row += 1
        d += timedelta(days=1)

    wb.save(filename)



# ═══════════════════════════════════════════════════════════════════════════════
# POLLAUS — laitteita minuutin välein, max 30 min
# ═══════════════════════════════════════════════════════════════════════════════

BEEP_SOUND = "/usr/share/sounds/freedesktop/stereo/complete.oga"

def beep() -> None:
    """Soittaa merkkiäänen, kun laitteelta on vastaanotettu tiedot.

    Ääni soitetaan PulseAudio/PipeWiren oletuslaitteesta (paplay) — jos
    langattomat kuulokkeet eivät ole päällä, ääni menee koneen kaiuttimiin.
    Varalla terminaalin kellomerkki. Epäonnistuminen ei keskeytä ohjelmaa.
    """
    import subprocess
    try:
        subprocess.run(["paplay", BEEP_SOUND], timeout=5,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        print("\a", end="", flush=True)


# Lukijafunktiot laiteavaimittain. DEVICE_TYPES (measurement.py) määrittää
# laitelistan; tämä sitoo kunkin laitteen sen BLE-lukijaan. Uusi laite
# lisätään rekisteröimällä se molempiin.
READERS = {
    "contour": read_contour,
    "omron":   read_omron,
    "beurer":  read_beurer,
}


async def collect_devices(macs: dict,
                          duration_minutes: int = 30,
                          interval_seconds: int = 60) -> dict:
    """Skannaa annettuja laitteita kunnes kaikilta on saatu tiedot tai aika loppuu.

    Käydään pollauskierroksia: jokaisella kierroksella yritetään lukea jokaista
    hakulistalla vielä olevaa laitetta. Kun joltain laitteelta saadaan tiedot,
    se poistetaan hakulistalta. Kierrosten välillä odotetaan kunnes minuutti
    on kulunut kierroksen alusta. Lopetetaan kun kaikki valmiita tai
    `duration_minutes` minuuttia on kulunut.

    macs: {laiteavain: MAC|None} — vain ne avaimet, joilla on MAC ja jotka
          löytyvät READERS-rekisteristä, ovat hakulistalla.
    Palauttaa sanakirjan {laiteavain: [lukijan tuottamia sanakirjoja]}.
    """
    pending = {name: mac for name, mac in macs.items()
               if mac and name in READERS}
    results = {name: [] for name in pending}
    if not pending:
        return results

    loop = asyncio.get_event_loop()
    deadline = loop.time() + duration_minutes * 60

    round_num = 0
    while pending and loop.time() < deadline:
        round_num += 1
        round_start = loop.time()
        remaining_min = (deadline - round_start) / 60
        print(f"\n{'═' * 70}")
        print(f"  Kierros {round_num}  ·  jäljellä {len(pending)} laitetta "
              f"({', '.join(pending.keys())})  ·  aikaa {remaining_min:.1f} min")
        print(f"{'═' * 70}")

        done_now = []
        for name in list(pending.keys()):
            if loop.time() >= deadline:
                break
            mac = pending[name]
            try:
                data = await READERS[name](mac)
            except Exception as e:
                print(f"[{name}] virhe: {e}")
                data = []
            if data:
                results[name] = data
                done_now.append(name)
                print(f"[{name}] ✓ tiedot saatu — poistetaan hakulistalta.")
                beep()

        for name in done_now:
            del pending[name]

        if not pending or loop.time() >= deadline:
            break

        elapsed = loop.time() - round_start
        wait = min(max(0.0, interval_seconds - elapsed),
                   max(0.0, deadline - loop.time()))
        if wait > 0:
            print(f"\n[Odotetaan {wait:.0f} s seuraavaan pollauskierrokseen…]")
            await asyncio.sleep(wait)

    if pending:
        print(f"\n⏱  Aikaraja {duration_minutes} min täynnä — ei saatu tietoja: "
              f"{', '.join(pending.keys())}")

    return results


# ═══════════════════════════════════════════════════════════════════════════════
# PÄÄOHJELMA
# ═══════════════════════════════════════════════════════════════════════════════

def load_config(config_file: str = "settings.ini") -> dict:
    """Lue laitelista (MAC-osoitteet) ja asetukset settings.ini-tiedostosta.

    Palauttaa sanakirjan:
        {
          "devices":   {laiteavain: MAC|None},  # kaikki DEVICE_TYPES-avaimet
          "directory": ...,
          "filename":  ...,
        }
    [devices]-osion avaimet luetaan geneerisesti laitelistan (DEVICE_TYPES)
    mukaan; tuntemattomista avaimista varoitetaan.
    """
    import configparser, os
    from datetime import datetime
    config = configparser.ConfigParser()

    defaults = {
        "devices":   {dt.key: None for dt in DEVICE_TYPES},
        "directory":  None,
        "filename":   "terveysmittaukset.xlsx",
        "period":     "year",
    }

    # Etsi asetustiedosto ohjelman hakemistosta tai nykyisestä hakemistosta
    script_dir = os.path.dirname(os.path.abspath(__file__))
    found_path = None
    for path in [config_file, os.path.join(script_dir, config_file)]:
        if os.path.exists(path):
            found_path = path
            break

    if not found_path:
        print(f"Huom: {config_file} ei löydy — käytetään oletusasetuksia.")
        return defaults

    config.read(found_path, encoding="utf-8")

    # Lue MAC-osoitteet laitelistan mukaan
    if "devices" in config:
        for key, raw in config["devices"].items():
            val = raw.strip()
            if not val:
                continue
            if key in defaults["devices"]:
                defaults["devices"][key] = val
            else:
                print(f"Huom: tuntematon laite '{key}' settings.ini:ssä — ohitetaan.")

    # Lue asetukset
    if "settings" in config:
        now = datetime.now()
        directory = config["settings"].get("directory", "").strip()
        filename  = config["settings"].get("filename",  "terveysmittaukset.xlsx").strip()

        # Kuukausitiedosto, jos nimimallissa käytetään {month}-muuttujaa;
        # muuten vuositiedosto. Ohjaa uuden tiedoston päivärivien laajuutta.
        defaults["period"] = "month" if "{month}" in filename else "year"

        # Korvaa päivämäärämuuttujat
        filename = filename.replace("{year}",  now.strftime("%Y"))
        filename = filename.replace("{month}", now.strftime("%m"))
        filename = filename.replace("{day}",   now.strftime("%d"))

        if directory:
            os.makedirs(directory, exist_ok=True)
            defaults["filename"] = os.path.join(directory, filename)
        else:
            defaults["filename"] = filename

        defaults["directory"] = directory or script_dir

    return defaults




if __name__ == "__main__":
    cfg = load_config()                     # Lue MAC-osoitteet asetustiedostosta

    devices_lines = "\n".join(
        f"  {dt.key:<8}= {cfg['devices'].get(dt.key) or 'ei asetettu'}   ({dt.label})"
        for dt in DEVICE_TYPES)
    device_flags = "  ".join(f"--{dt.key}" for dt in DEVICE_TYPES)
    parser = argparse.ArgumentParser(
        description="Terveysmittaukset BLE-lukija (laitelista settings.ini:stä)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
MAC-osoitteet luetaan tiedostosta settings.ini (sama hakemisto kuin ohjelma).
Tällä hetkellä asetustiedostossa:
{devices_lines}

Esimerkkejä:
  python3 health.py --scan
  python3 health.py --contour        (käyttää settings.ini osoitetta)
  python3 health.py --kaikki         (lukee kaikki laitelistan laitteet)
  python3 health.py --contour AA:BB:CC:DD:EE:FF  (ylikirjoittaa ini-osoitteen)
  python3 health.py --kaikki --excel toukokuu.xlsx
        """
    )
    parser.add_argument("--scan", action="store_true", help="Etsi BLE-laitteita lähistöltä")
    # CLI-liput luodaan laitelistan (DEVICE_TYPES) mukaan: yksi --<avain> per laite.
    for dt in DEVICE_TYPES:
        parser.add_argument(f"--{dt.key}", nargs="?", const=cfg["devices"].get(dt.key),
                            metavar="MAC",
                            help=f"Lue {dt.label} (MAC settings.ini:stä tai anna itse)")
    parser.add_argument("--kaikki",   action="store_true", help="Lue kaikki laitelistan laitteet")
    parser.add_argument("--excel",    type=str, default=None, help=f"Tallennustiedosto (oletus: {cfg['filename']})")
    parser.add_argument("--no-excel", action="store_true", help="Näytä tulokset vain terminaalissa")
    args = parser.parse_args()

    if args.scan:
        asyncio.run(scan_devices()); sys.exit(0)

    # Kokoa pollattavien laitteiden MAC-osoitteet laitelistan mukaan.
    selected = {}
    for dt in DEVICE_TYPES:
        mac = getattr(args, dt.key)
        if args.kaikki:                                                 # --kaikki täydentää ini-tiedoston osoitteilla
            mac = mac or cfg["devices"].get(dt.key)
        selected[dt.key] = mac

    if not any(selected.values()):
        parser.print_help(); sys.exit(0)

    excel_file = args.excel or cfg["filename"]                          # Tulostiedoston polku settings.ini:stä (sisältää esim. vuoden).

    device_results = asyncio.run(                                       # Pollataan kaikkia annettuja laitteita minuutin välein max 30 min.
        collect_devices(selected)
    )

    if not any(device_results.values()):
        print("\nEi mittauksia saatu yhdeltäkään laitteelta."); sys.exit(1)

    if not args.no_excel:
        # Laite voi palauttaa useita lukemia (esim. Contour ~20 verensokeria).
        # Kootaan kaikki päivätasolle: yksi Measurement per kalenteripäivä.
        measurements = combine_measurements_by_day(device_results)

        if not os.path.exists(excel_file):                              # Luo päiväpohjainen tiedosto, jos sitä ei vielä ole.
            create_new_file(excel_file, period=cfg["period"])
            print(f"Luotu uusi päiväpohjainen tiedosto: {excel_file}")

        print(f"\nKirjoitetaan {len(measurements)} mittauspäivää tiedostoon {excel_file}:")
        for m in measurements:
            day = m.timestamp.strftime("%d.%m.%Y") if m.timestamp else "?"
            vals = ", ".join(
                f"{header}={value}"
                for (header, _), value in zip(Measurement.COLUMNS[1:], m.to_row()[1:])
                if value not in (None, ""))
            print(f"  {day}: {vals or '-'}")

        written, unmatched = write_measurements_by_date(excel_file, measurements)  # Etsii oikean päivärivin ja täyttää sen.
        print(f"\n{written} päivän mittaukset tallennettu tiedostoon {excel_file}.")
        if unmatched:
            print(f"⚠  {len(unmatched)} mittauspäivää ei voitu sijoittaa — "
                  f"päivää ei löytynyt taulukosta (onko vuosi/kuukausi oikea?):")
            for m in unmatched:
                print(f"     {m.timestamp.strftime('%d.%m.%Y') if m.timestamp else '?'}")
