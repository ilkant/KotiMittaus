#!/usr/bin/env python3
"""
testaus.py — Measurement-luokan ja Excel-tallennuksen testi

Luo yhden Measurement-olion keksityillä mittausarvoilla ja kirjoittaa sen
tämän päivän aikaleimalla settings.ini-tiedostossa määriteltyyn tulostiedostoon
"Mittaukset"-välilehdelle. Jos tiedosto on jo olemassa, rivi lisätään
olemassa olevien jatkoksi. Jos välilehteä ei vielä ole, se luodaan otsikoineen.

Käyttö:
    python3 testaus.py
"""

import os, sys
from datetime import datetime
from openpyxl import Workbook

from measurement import Measurement, save_measurements_sheet, append_measurement
from health import load_config


def create_new_file(filename: str) -> None:
    """Luo uuden Excel-tiedoston, jossa on 'Mittaukset'-välilehti otsikoineen.

    Käytetään silloin, kun tulostiedostoa ei vielä ole olemassa. Tarvittavat
    hakemistot luodaan automaattisesti.
    """
    target_dir = os.path.dirname(filename)
    if target_dir:
        os.makedirs(target_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    save_measurements_sheet(wb, [])  # luo välilehden otsikkorivillä
    # save_measurements_sheet kirjoittaa tyhjään tauluun "Ei mittauksia." -solun;
    # poistetaan se, jotta ensimmäinen varsinainen rivi voidaan kirjoittaa rivi 2:lle.
    ws = wb["Mittaukset"]
    if ws.cell(2, 1).value == "Ei mittauksia.":
        ws.cell(2, 1).value = None
    wb.save(filename)


if __name__ == "__main__":
    cfg = load_config()
    filename = cfg["tiedosto"]

    # Keksityt mittausarvot — tämän päivän aikaleimalla
    m = Measurement( timestamp = datetime.now(), glucose_mmol=5.8, weight=78.4, systolic=128, diastolic=82, pulse=65,
                     fat=22.1, water=56.3, muscle=38.2, bone=3.1, bmr=1720, amr=2580,
    )

    print(f"Tulostiedosto:  {filename}")
    print(f"Aikaleima:      {m.timestamp.strftime('%d.%m.%Y %H:%M')}")
    print(f"Kirjoitettava rivi:")
    for (header, _), value in zip(Measurement.COLUMNS, m.to_row()):
        print(f"  {header:<28} {value}")

    # Luo tiedosto otsikkoineen, jos sitä ei ole vielä olemassa
    if not os.path.exists(filename):
        create_new_file(filename)
        print(f"\nUusi tiedosto luotu otsikkorivillä: {filename}")

    append_measurement(filename, m)
    print(f"\nRivi tallennettu onnistuneesti tiedostoon {filename}.")
