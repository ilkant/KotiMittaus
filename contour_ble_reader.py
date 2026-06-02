#!/usr/bin/env python3
"""
Contour Care 7951H verensokerimittarin BLE-lukija
Käyttää Bluetooth GATT Glucose Service -standardia (UUID 0x1808)

Asennus:
    pip install bleak openpyxl

Käyttö:
    # Vaihe 1: etsi mittarin MAC-osoite
    python3 contour_ble_reader.py --scan

    # Vaihe 2: lue mittaukset (korvaa MAC-osoite)
    python3 contour_ble_reader.py --mac AA:BB:CC:DD:EE:FF

    # Vaihe 3: tallenna LibreOfficeen
    python3 contour_ble_reader.py --mac AA:BB:CC:DD:EE:FF --excel verensokeri.xlsx

Huomio: Mittarin Bluetooth täytyy olla päällä (paina BT-nappia mittarissa).
"""

import asyncio, struct, argparse, sys
from datetime import datetime, timedelta

try:
    from bleak import BleakScanner, BleakClient
    from bleak.exc import BleakError
except ImportError:
    print("Virhe: bleak-kirjasto puuttuu. Asenna: pip install bleak")
    sys.exit(1)

# Bluetooth GATT UUID:t verensokerimittareille
GLUCOSE_SERVICE_UUID        = "00001808-0000-1000-8000-00805f9b34fb"
GLUCOSE_MEASUREMENT_UUID    = "00002a18-0000-1000-8000-00805f9b34fb"
GLUCOSE_CONTEXT_UUID        = "00002a34-0000-1000-8000-00805f9b34fb"
RACP_UUID                   = "00002a52-0000-1000-8000-00805f9b34fb"  # Record Access Control Point

# RACP-komennot (Bluetooth SIG standardi)
RACP_REPORT_ALL_RECORDS     = bytes([0x01, 0x01])  # Hae kaikki mittaukset
RACP_RESPONSE_SUCCESS       = 0x06

MEAL_MARKERS = {
    0x00: "",
    0x01: "Paasto",
    0x02: "Ennen ateriaa",
    0x03: "Aterian jälkeen",
    0x04: "Satunnainen",
    0x05: "Yö",
    0x06: "Aamupala",
    0x07: "Lounas",
    0x08: "Illallinen",
}

measurements = []
racp_done = asyncio.Event()


def parse_glucose_measurement(data: bytes) -> dict | None:
    """Purkaa GATT 2A18 Glucose Measurement -paketin."""
    if len(data) < 7:
        return None

    flags = data[0]
    seq   = struct.unpack_from("<H", data, 1)[0]

    # Päivämäärä ja aika (vuosi=2B, kk=1B, pv=1B, h=1B, min=1B, s=1B)
    year, month, day, hour, minute, second = struct.unpack_from("<HBBBBB", data, 3)
    try:
        timestamp = datetime(year, month, day, hour, minute, second)
    except ValueError:
        timestamp = None

    offset = 10

    # Aikasiirtymä minuutteina (jos lippu 0 asetettu)
    if flags & 0x01:
        time_offset = struct.unpack_from("<h", data, offset)[0]
        offset += 2
        if timestamp:
            timestamp += timedelta(minutes=time_offset)

    # Glukoosiarvo (SFLOAT, 2 tavua) — lippu 1 = arvo mukana
    glucose_mmol = None
    if flags & 0x02:
        raw = struct.unpack_from("<H", data, offset)[0]
        offset += 2
        mantissa = raw & 0x0FFF
        if mantissa & 0x0800:
            mantissa -= 0x1000
        exponent = (raw >> 12) & 0x0F
        if exponent & 0x08:
            exponent -= 0x10
        value = mantissa * (10 ** exponent)

        # Lippu 2: yksikkö — 0=kg/L, 1=mol/L
        if flags & 0x04:
            glucose_mmol = value * 1000  # mol/L → mmol/L
        else:
            glucose_mmol = value * 1000 * 18.0182 / 10  # kg/L → mmol/L (approksimaatio)
            # Contour raportoi usein suoraan mmol/L mol/L -muodossa
            if glucose_mmol > 40:
                glucose_mmol = value * 1000  # käytä sellaisenaan jos arvo järjetön

    return {
        "seq": seq,
        "timestamp": timestamp,
        "glucose_mmol": round(glucose_mmol, 1) if glucose_mmol else None,
    }


def glucose_notification(sender, data: bytearray):
    result = parse_glucose_measurement(bytes(data))
    if result and result["glucose_mmol"]:
        measurements.append(result)
        ts = result["timestamp"].strftime("%d.%m.%Y %H:%M") if result["timestamp"] else "?"
        print(f"  [{result['seq']:4d}] {ts}  →  {result['glucose_mmol']:.1f} mmol/L")


def context_notification(sender, data: bytearray):
    pass  # Kontekstitiedot (ateria ym.) — ei käytetä tässä versiossa


def racp_notification(sender, data: bytearray):
    # Tarkistetaan onko siirto valmis
    if len(data) >= 4 and data[0] == 0x06:
        racp_done.set()


async def read_meter(mac_address: str):
    """Yhdistä mittariin ja lue kaikki tallennetut mittaukset."""
    print(f"\nYhdistetään mittariin {mac_address} ...")
    print("Varmista että mittarin Bluetooth on päällä!\n")

    async with BleakClient(mac_address, timeout=20.0) as client:
        if not client.is_connected:
            print("Virhe: yhteys epäonnistui")
            return []

        print("Yhdistetty! Luetaan mittauksia...\n")

        # Tilataan notifikaatiot
        await client.start_notify(GLUCOSE_MEASUREMENT_UUID, glucose_notification)
        await client.start_notify(RACP_UUID, racp_notification)

        try:
            await client.start_notify(GLUCOSE_CONTEXT_UUID, context_notification)
        except Exception:
            pass  # Kaikki mittarit eivät tue kontekstiominaisuutta

        # Pyydä kaikki tallennetut mittaukset
        await client.write_gatt_char(RACP_UUID, RACP_REPORT_ALL_RECORDS, response=True)

        # Odota enintään 30 sekuntia
        try:
            await asyncio.wait_for(racp_done.wait(), timeout=30.0)
        except asyncio.TimeoutError:
            print("(Aikakatkaisu — käytetään jo saadut mittaukset)")

        await client.stop_notify(GLUCOSE_MEASUREMENT_UUID)
        await client.stop_notify(RACP_UUID)

    return measurements


async def scan_devices():
    """Etsi BLE-laitteita lähistöltä."""
    print("Etsitään Bluetooth-laitteita 10 sekuntia...")
    print("Paina mittarin BT-nappia nyt!\n")

    devices = await BleakScanner.discover(timeout=10.0)

    contour_devices = []
    other_devices = []

    for d in devices:
        name = d.name or ""
        if any(x in name.upper() for x in ["CONTOUR", "BAYER", "ASCENSIA", "BGM"]):
            contour_devices.append(d)
        else:
            other_devices.append(d)

    if contour_devices:
        print("=== Löydetyt Contour-laitteet ===")
        for d in contour_devices:
            print(f"  Nimi: {d.name or 'tuntematon':30s}  MAC: {d.address}")
    else:
        print("Contour-laitteita ei löydy automaattisesti.")
        print("Kaikki löydetyt laitteet:")
        for d in sorted(other_devices, key=lambda x: x.rssi or -999, reverse=True)[:15]:
            print(f"  {d.address}  RSSI: {d.rssi or '?':4}  {d.name or ''}")

    print("\nKopioi MAC-osoite ja käytä --mac -valitsinta.")


def save_to_excel(data: list, filename: str):
    """Tallentaa mittaukset Excel/LibreOffice-tiedostoon."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Virhe: openpyxl puuttuu. Asenna: pip install openpyxl")
        return

    if not data:
        print("Ei tallennettavia mittauksia.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Verensokeri"

    # Otsikkorivi
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Järj.nro", "Päivämäärä", "Kellonaika", "Verensokeri (mmol/L)", "Huomio"]
    col_widths = [10, 15, 12, 22, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    # Viiterajat mmol/L
    LOW_LIMIT  = 4.0
    HIGH_LIMIT = 8.0

    low_fill  = PatternFill("solid", start_color="FFE0E0")  # vaaleanpunainen = matala
    high_fill = PatternFill("solid", start_color="FFEAC0")  # oranssi = korkea
    ok_fill   = PatternFill("solid", start_color="E8F5E9")  # vihreä = OK

    data_sorted = sorted(data, key=lambda x: x["timestamp"] or datetime.min)

    for row_idx, m in enumerate(data_sorted, 2):
        ts = m["timestamp"]
        glucose = m["glucose_mmol"]

        ws.cell(row=row_idx, column=1, value=m["seq"])
        ws.cell(row=row_idx, column=2, value=ts.strftime("%d.%m.%Y") if ts else "")
        ws.cell(row=row_idx, column=3, value=ts.strftime("%H:%M") if ts else "")

        g_cell = ws.cell(row=row_idx, column=4, value=glucose)
        g_cell.number_format = "0.0"
        g_cell.alignment = Alignment(horizontal="center")

        note = ""
        if glucose is not None:
            if glucose < LOW_LIMIT:
                note = "⚠ Matala"
                row_fill = low_fill
            elif glucose > HIGH_LIMIT:
                note = "▲ Korkea"
                row_fill = high_fill
            else:
                note = "✓ OK"
                row_fill = ok_fill
        else:
            row_fill = None

        ws.cell(row=row_idx, column=5, value=note)

        if row_fill:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = row_fill

        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = border

    # Yhteenveto-välilehti
    ws2 = wb.create_sheet("Yhteenveto")
    n = len(data_sorted)
    data_row_end = n + 1  # viimeinen datarivi Verensokeri-välilehdellä

    ws2["A1"] = "Yhteenveto"
    ws2["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ("Mittauksia yhteensä", f"=COUNTA(Verensokeri!D2:D{data_row_end})"),
        ("Keskiarvo (mmol/L)",  f"=ROUND(AVERAGE(Verensokeri!D2:D{data_row_end}),1)"),
        ("Minimi (mmol/L)",     f"=MIN(Verensokeri!D2:D{data_row_end})"),
        ("Maksimi (mmol/L)",    f"=MAX(Verensokeri!D2:D{data_row_end})"),
        ("Matalia (<4.0)",      f'=COUNTIF(Verensokeri!D2:D{data_row_end},"<4")'),
        ("Korkeita (>8.0)",     f'=COUNTIF(Verensokeri!D2:D{data_row_end},">8")'),
        ("Tavoitealueella",     f'=COUNTIFS(Verensokeri!D2:D{data_row_end},">=4",Verensokeri!D2:D{data_row_end},"<=8")'),
    ]

    for r, (label, formula) in enumerate(summary_rows, 3):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=formula)
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 15

    # Viivakuvaaja
    chart = LineChart()
    chart.title = "Verensokeri aikajärjestyksessä"
    chart.style = 10
    chart.y_axis.title = "mmol/L"
    chart.x_axis.title = "Mittaus"
    chart.height = 12
    chart.width  = 22

    glucose_ref = Reference(ws, min_col=4, min_row=1, max_row=n + 1)
    chart.add_data(glucose_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = "2E75B6"

    ws2.add_chart(chart, "A12")

    wb.save(filename)
    print(f"\nTiedosto tallennettu: {filename}")
    print(f"Avaa LibreOffice Calcissa: libreoffice --calc {filename}")


def main():
    parser = argparse.ArgumentParser(
        description="Contour Care 7951H BLE-lukija"
    )
    parser.add_argument("--scan", action="store_true",
                        help="Etsi BLE-laitteita lähistöltä")
    parser.add_argument("--mac", type=str,
                        help="Mittarin Bluetooth MAC-osoite (esim. AA:BB:CC:DD:EE:FF)")
    parser.add_argument("--excel", type=str, default="verensokeri.xlsx",
                        help="Tallennustiedoston nimi (oletus: verensokeri.xlsx)")
    parser.add_argument("--no-excel", action="store_true",
                        help="Älä tallenna Exceliin, näytä vain terminaalissa")
    args = parser.parse_args()

    if args.scan:
        asyncio.run(scan_devices())
        return

    if not args.mac:
        parser.print_help()
        print("\nEsimerkki:\n  python3 contour_ble_reader.py --scan")
        print("  python3 contour_ble_reader.py --mac AA:BB:CC:DD:EE:FF")
        sys.exit(1)

    results = asyncio.run(read_meter(args.mac))

    if not results:
        print("\nEi mittauksia saatu. Tarkista:")
        print("  1. Mittarin Bluetooth on päällä (paina BT-nappia)")
        print("  2. MAC-osoite on oikein (käytä --scan)")
        print("  3. Mittari ei ole yhdistynyt puhelimeen samaan aikaan")
        sys.exit(1)

    print(f"\nSaatiin {len(results)} mittausta.")

    if not args.no_excel:
        save_to_excel(results, args.excel)


if __name__ == "__main__":
    main()
